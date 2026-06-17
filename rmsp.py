from __future__ import annotations

import json
import logging
from dataclasses import dataclass
from io import BytesIO
from typing import Any
from urllib.parse import urlencode

import aiohttp

from services.net import ipv4_connector
from openpyxl import load_workbook

from services.inn_utils import normalize_inn

logger = logging.getLogger("crnabot.rmsp")

RMSP_BASE = "https://rmsp.nalog.ru"
REPORT_URL = f"{RMSP_BASE}/report.xlsx"

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
)

TIMEOUT = aiohttp.ClientTimeout(total=60)

CATEGORY_BY_CODE = {
    "1": "Микропредприятие",
    "2": "Малое предприятие",
    "3": "Среднее предприятие",
}


class RmspError(Exception):
    pass


@dataclass
class RmspRecord:
    inn: str
    name: str
    ogrn: str
    category: str | None
    in_registry: bool
    region: str | None = None


def _pick(row: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return str(value).strip()
    return ""


def _inn_list_search_params(inn: str) -> dict[str, str]:
    return {
        "mode": "inn-list",
        "page": "1",
        "innList": normalize_inn(inn),
        "pageSize": "10",
        "sortField": "NAME_EX",
        "sort": "ASC",
    }


def _parse_category(row: dict[str, Any]) -> str | None:
    if row.get("is_active") in {0, "0", False}:
        out_date = _pick(row, "dtregistryout", "dtRegistryOut")
        if out_date:
            return f"Исключён из реестра ({out_date.split()[0]})"
        return "Исключён из реестра"

    raw_category = row.get("category")
    if raw_category not in (None, ""):
        code = str(int(raw_category)) if str(raw_category).isdigit() else str(raw_category).strip()
        if code in CATEGORY_BY_CODE:
            return CATEGORY_BY_CODE[code]

    for key in (
        "catTypeName",
        "categoryName",
        "catName",
        "CAT_NAME",
        "cat_type_name",
        "mspCategory",
    ):
        value = _pick(row, key)
        if value:
            return value

    code = _pick(row, "catType", "cat", "CAT", "mspCat", "categoryCode", "cat_code")
    if code in CATEGORY_BY_CODE:
        return CATEGORY_BY_CODE[code]
    if code.isdigit() and int(code) in {1, 2, 3}:
        return CATEGORY_BY_CODE[str(int(code))]
    return None


def _rows_from_payload(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [x for x in payload if isinstance(x, dict)]
    if not isinstance(payload, dict):
        return []

    for key in ("data", "rows", "items", "result", "content"):
        bucket = payload.get(key)
        if isinstance(bucket, list):
            return [x for x in bucket if isinstance(x, dict)]
        if isinstance(bucket, dict):
            for nested_key in ("rows", "data", "items"):
                nested = bucket.get(nested_key)
                if isinstance(nested, list):
                    return [x for x in nested if isinstance(x, dict)]
    return []


def _match_row(rows: list[dict[str, Any]], inn: str) -> dict[str, Any] | None:
    normalized = normalize_inn(inn)
    for row in rows:
        row_inn = normalize_inn(_pick(row, "inn", "INN", "innUl", "innIp"))
        if row_inn and row_inn == normalized:
            return row
    return None


def _category_from_xlsx(data: bytes, inn: str) -> str | None:
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return None

    normalized = normalize_inn(inn)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            header_text = [str(cell or "").strip().lower() for cell in header]
            category_idx = next(
                (i for i, text in enumerate(header_text) if "категор" in text),
                None,
            )
            inn_idx = next(
                (i for i, text in enumerate(header_text) if text in {"инн", "инн юл", "инн ип"}),
                None,
            )
            for row in rows:
                if not row:
                    continue
                if inn_idx is not None and len(row) > inn_idx:
                    row_inn = normalize_inn(row[inn_idx])
                    if row_inn != normalized:
                        continue
                if category_idx is not None and len(row) > category_idx:
                    value = str(row[category_idx] or "").strip()
                    if value:
                        return value
                for cell in row:
                    text = str(cell or "").strip()
                    if text in CATEGORY_BY_CODE.values():
                        return text
    finally:
        workbook.close()
    return None


def _base_headers(*, json_request: bool = False) -> dict[str, str]:
    headers = {
        "User-Agent": USER_AGENT,
        "Accept-Language": "ru-RU,ru;q=0.9",
        "Referer": f"{RMSP_BASE}/",
        "Origin": RMSP_BASE,
        "Accept-Encoding": "gzip, deflate",
    }
    if json_request:
        headers.update(
            {
                "Accept": "application/json, text/javascript, */*; q=0.01",
                "X-Requested-With": "XMLHttpRequest",
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            }
        )
    return headers


def _is_xlsx(data: bytes) -> bool:
    return data[:2] == b"PK"


async def _warmup(session: aiohttp.ClientSession) -> None:
    async with session.get(
        f"{RMSP_BASE}/",
        headers=_base_headers(),
    ) as resp:
        await resp.read()


async def _post_form(
    session: aiohttp.ClientSession,
    path: str,
    data: dict[str, str],
    *,
    json_response: bool = True,
) -> Any:
    async with session.post(
        f"{RMSP_BASE}{path}",
        data=urlencode(data),
        headers=_base_headers(json_request=True),
    ) as resp:
        body = await resp.read()
        if resp.status != 200:
            raise RmspError(f"HTTP {resp.status}")
        if not json_response:
            return body
        text = body.decode("utf-8", errors="replace")
        if not text.strip():
            return {}
        return json.loads(text)


async def _search_row(session: aiohttp.ClientSession, inn: str) -> tuple[dict[str, Any] | None, dict[str, str]]:
    search_params = _inn_list_search_params(inn)
    queries = (
        search_params,
        {"mode": "search", "query": inn, "page": "1", "pageSize": "10"},
        {"mode": "search-ul", "queryUl": inn, "page": "1", "pageSize": "10"},
        {"mode": "search", "queryAll": inn, "page": "1", "pageSize": "10"},
    )
    for data in queries:
        try:
            payload = await _post_form(session, "/search-proc.json", data)
            rows = _rows_from_payload(payload)
            row = _match_row(rows, inn)
            if row:
                return row, data
            if isinstance(payload, dict) and payload.get("rowCount"):
                logger.info(
                    "Реестр МСП search %s rowCount=%s без точного совпадения inn=%s",
                    data.get("mode"),
                    payload.get("rowCount"),
                    inn,
                )
        except Exception as exc:
            logger.info("Реестр МСП search %s failed: %s", data.get("mode"), exc)
    return None, search_params


async def _download_report(session: aiohttp.ClientSession, search_params: dict[str, str]) -> bytes:
    body = urlencode(search_params)
    headers = _base_headers()
    headers.update(
        {
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "*/*",
        }
    )
    async with session.post(REPORT_URL, data=body, headers=headers) as resp:
        data = await resp.read()
        content_type = resp.headers.get("Content-Type", "")
        if resp.status != 200:
            raise RmspError(f"Не удалось скачать выписку из реестра МСП (HTTP {resp.status}).")
        if not _is_xlsx(data):
            preview = data[:240].decode("utf-8", errors="replace")
            logger.error(
                "Реестр МСП report is not xlsx inn-list=%s type=%s size=%s preview=%r",
                search_params.get("innList"),
                content_type,
                len(data),
                preview,
            )
            raise RmspError("Не удалось скачать выписку из реестра МСП.")
    return data


def _record_from_row(row: dict[str, Any], inn: str, category: str | None) -> RmspRecord:
    return RmspRecord(
        inn=inn,
        name=_pick(row, "name_ex", "name", "NAME", "nameUl", "nameEx", "shortName"),
        ogrn=_pick(row, "ogrn", "OGRN", "ogrnUl", "ogrnIp"),
        category=category,
        in_registry=True,
        region=_pick(row, "region", "regionName", "REGION", "regioncode"),
    )


def _session() -> aiohttp.ClientSession:
    return aiohttp.ClientSession(
        timeout=TIMEOUT,
        cookie_jar=aiohttp.CookieJar(unsafe=True),
        trust_env=False,
        connector=ipv4_connector(),
    )


async def lookup_rmsp(inn: str) -> RmspRecord:
    inn = normalize_inn(inn)
    async with _session() as session:
        await _warmup(session)
        row, _search_params = await _search_row(session, inn)
        if not row:
            return RmspRecord(
                inn=inn,
                name="",
                ogrn="",
                category="Не найден в реестре МСП",
                in_registry=False,
            )
        return _record_from_row(row, inn, _parse_category(row))


async def download_rmsp_report(inn: str) -> tuple[bytes, RmspRecord]:
    inn = normalize_inn(inn)
    search_params = _inn_list_search_params(inn)
    async with _session() as session:
        await _warmup(session)
        payload = await _post_form(session, "/search-proc.json", search_params)
        row = _match_row(_rows_from_payload(payload), inn)
        if not row:
            raise RmspError(f"Организация с ИНН {inn} не найдена в реестре МСП.")

        report = await _download_report(session, search_params)
        category = _parse_category(row) or _category_from_xlsx(report, inn)
        return report, _record_from_row(row, inn, category)


download_rmsp_pdf = download_rmsp_report
